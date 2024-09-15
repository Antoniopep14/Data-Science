#pipenv install openpyxl   para instalar la dependencia de excel
import openpyxl
wb = openpyxl.load_workbook("planilla.xlsx")
#metodos
print(wb.sheetnames)#['Hoja1'] da los nombres de las hojas dentro del archivo
print(wb["Hoja1"])#<Worksheet "Hoja1">
#esta segunda es una forma de acceder a las hojas
#o tambien podemos pedir que nos devuelva la hoja que se encuentra activa
hoja = wb.active
#en este caso nos va a devolver la primera por defecto a menos que la cambiemos
print(hoja)#<Worksheet "Hoja1">
#ahira agregaremos una nueva hoja
wb.create_sheet("Hoja3")#esto la crea, pero aun no la guarda en nuestro archivo
print(wb.sheetnames)#['Hoja1', 'Hoja3']
#ahora abriremos la hoja 3
hoja3 = wb["Hoja3"]
#y le cambiaremos el titulo
hoja3.title = "Nuevo hoja3"
#ahora veamos la cantidad de celdas y filas utilizadas
# print(
#     hoja.max_row,
#     hoja.max_column,
# )#4 3
#ahora seleccionemos una celda
celda = hoja["A1"]
print(celda)#<Cell 'Hoja1'.A1>
print(celda.value)#Nombre
#si quisieramos cambiarle el valor a esa celda seria
celda.value = "Nombre completo"

#ahora accedamos a la celda de una forma mas origramable
celda2 = hoja.cell(row=2, column=1)
#recordar que las las filas y columnas parten de 1, no de 0
print(celda2.value,
      celda2.row,
      celda2.column,
      celda2.coordinate)#Nicolas 2 1 A2

#con esta informacion ya podemos acceder derechamente a toda la informacion
#que contienen todas las celdas y todas la columnas
#iterandolas con un for

for fila in range(1, hoja.max_row + 1):
    for columna in range(1, hoja.max_column + 1):
        celdax = hoja.cell(row=fila, column=columna)
        #print(celdax.value)#Nmobre completo|Edad|Correo|Nicolas|25
#nicolas@chanchito.com|Felipe|34|felipe@chanchito.com|Chanchito
#12|yo@chanchito.com
#ahora vamos a embellecerlo un poco
        print(fila, columna, celdax.value)

#ahora veamos como obtener una columna o fila en particular
columna = hoja["A"]
fila = hoja["1"]
print(columna)#(<Cell 'Hoja1'.A1>, <Cell 'Hoja1'.A2>, <Cell 'Hoja1'.A3>, <Cell 'Hoja1'.A4>)
print(fila)#(<Cell 'Hoja1'.A1>, <Cell 'Hoja1'.B1>, <Cell 'Hoja1'.C1>)


#ahora agreguemos elementos a las filas
hoja.append([1, 22, 333])#aqui debemos agregar como listado los valores que queremos agregar
print(hoja.rows)#<generator object Worksheet._cells_by_row at 0x0000023951D95240>
#para ver su contenido tendriamos que iterarlo con un for y listo

#para eliminar columnas o filas usamos:
#el primero es para filas, el segundo para columnas, y en el ()
#va el indice desde donde quiero empezar y el segundo
#es la cantidad de filas o columas que eliminara
#sino le pasamos nada al segundo su valor por defecto es 1

#hoja.delete_cols(1, 1)
hoja.delete_rows(1, 1)

#ahora guardemos los cambios
wb.save("plantilla2.xlsx")#tenemos que indicar el nombre del archivo, ademas la ruta
#donde lo guardaremos, sino la indicamos la guarda donde ejecutamos



